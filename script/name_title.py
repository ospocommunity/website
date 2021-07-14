# Put the ApacheConAsia讲师信息收集表.xlsx、ApacheConSessions.xlsx file in the project root directory
# will be generated in the root directory zh_name_title.csv and en_name_title.csv file

from pypinyin import lazy_pinyin
import openpyxl
import csv
from translate import Translator
 
# First, define a CSV file
zh_head =["zh_name","zh_title","track","zh_position"]
en_head =["en_name","en_title","track","en_position"]
with open("./zh_name_title.csv","a+",encoding="utf-8",newline="") as zh_f:
   csvf=csv.writer(zh_f)
   csvf.writerow(zh_head)
with open("./en_name_title.csv","a+",encoding="utf-8",newline="") as en_f:
   csvf=csv.writer(en_f)
   csvf.writerow(en_head)

# Open the EXECL file
session=openpyxl.load_workbook("./ApacheConSessions.xlsx")
collect=openpyxl.load_workbook("./ApacheConAsia讲师信息收集表.xlsx")

# Take the string between two fixed symbols
def get_str_btw(str,begin,end):
   par=str.partition(begin)
   return (par[2].partition(end))[0][:]

# Translate Chinese name into English name
def translateName_ch_en(zh_name):
   name_list=lazy_pinyin(zh_name)
   xin=name_list[0]
   ming_list=name_list[1:]
   ming=""
   for i in ming_list:
      ming=ming+i
      en_name=ming.capitalize()+" "+xin.capitalize()
   return (en_name)

# Gets the active table object
session_active=session.active
collect_active=collect.active
flag=False
for cell1 in session_active['J']:
   str=cell1.value
   row=cell1.row
   mail_list=str.split(",")
   for i in mail_list:
      if row>1 :
         session_mail=get_str_btw(i,'<','>')  
         names=i.split(" <")
         zh_title=session_active.cell(row,7).value
         en_title=session_active.cell(row,6).value
         track=session_active.cell(row,1).value
         zh_name=names[0].strip()
         # For names with brackets, remove the brackets and the contents in the brackets
         zh_name=zh_name.split('(',1)[0]
         zh_name=zh_name.split('（',1)[0]
         if zh_title!=None and en_title!=None :
            # Judge whether it is a Chinese name
            if '\u4e00' <=zh_name<='\u9fff':
               # Translate Chinese name into English name
               en_name=translateName_ch_en(zh_name)
            else:
               en_name=zh_name
               for mail in collect_active['J'] :
                  if mail.value==session_mail :
                     zh_name=collect_active.cell(mail.row,3).value
                     break
            
            # According to the topic of Chinese speech, get the position information
            flag=False
            for position in collect_active['D']:
               if position.value==zh_title or position.value==en_title:
                  zh_position=collect_active.cell(position.row,6).value

                  # Translation from Chinese to English
                  translator=Translator(from_lang="chinese",to_lang="english")
                  en_position=translator.translate(zh_position)
                  flag=True
                  break
            if flag==False :
               zh_position=0
               en_position=0
            
            # Write data to a CSV file
            zh_data=[(zh_name,zh_title,track,zh_position)]
            en_data=[(en_name,en_title,track,en_position)]
            with open("./zh_name_title.csv","a+",encoding="utf-8",newline="") as zh_f:
               csvf=csv.writer(zh_f)
               csvf.writerows(zh_data)

            with open("./en_name_title.csv","a+",encoding="utf-8",newline="") as en_f:
               csvf=csv.writer(en_f)
               csvf.writerows(en_data)
   