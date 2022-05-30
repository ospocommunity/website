## ---------------------------------------------------------------------------
## Licensed to the Apache Software Foundation (ASF) under one or more
## contributor license agreements.  See the NOTICE file distributed with
## this work for additional information regarding copyright ownership.
## The ASF licenses this file to You under the Apache License, Version 2.0
## (the "License"); you may not use this file except in compliance with
## the License.  You may obtain a copy of the License at
##
##      http://www.apache.org/licenses/LICENSE-2.0
##
## Unless required by applicable law or agreed to in writing, software
## distributed under the License is distributed on an "AS IS" BASIS,
## WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
## See the License for the specific language governing permissions and
## limitations under the License.
## ---------------------------------------------------------------------------
from openpyxl import load_workbook
from PIL import Image


wb = load_workbook(r'ApacheConSessions.xlsx')
ws = wb[wb.sheetnames[0]]
save_path = r'../themes/apachecon/static/images/speaker'
file_name_arr = []
session_success_status = "审核通过"

try:
    for image in ws._images:
        print(image)
        # print(image.anchor._from)
        anchor_marker = image.anchor._from
        row = anchor_marker.row
        status = ws.cell(row+1, 2).value
        print(status)
        if status == session_success_status:
            file_name = str(row - 1 + 1000)

            img_index = 2

            tmp_file_name = file_name
            while tmp_file_name in file_name_arr:
                tmp_file_name = file_name + "_" + str(img_index)
                img_index += 1
            file_name = tmp_file_name

            file_name_arr.append(file_name)
            img = Image.open(image.ref).convert("RGB")

            (x, y) = img.size
            x_1 = 400
            y_1 = int(y * x_1 / x)
            out = img.resize((x_1, y_1), Image.ANTIALIAS)
            out.save('{0}\\{1}'.format(save_path, file_name + ".png"), quality=95)
    wb.close()
except Exception as e:
        print(e)
        wb.close()

