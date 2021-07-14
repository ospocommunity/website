# Put the ApacheConAsia讲师信息收集表.xlsx、ApacheConSessions.xlsx、drawing1.xml file in the project root directory
# will be generated in the root directory mapping.csv file

from xml.dom.minidom import parse
from pypinyin import lazy_pinyin
import xml.dom.minidom
import openpyxl
import csv

 
 # First, define a CSV file
head =["collect_row","pic","zh_name","en_name","position","track","title","mail","sessions_row"]
with open("./mapping.csv","a+",encoding="utf-8",newline="") as f:
   csvf=csv.writer(f)
   csvf.writerow(head)

# Using minidom parser to open XML document
DOMTree = xml.dom.minidom.parse("./drawing1.xml")

# Open the EXECL file
collect=openpyxl.load_workbook("./ApacheConAsia讲师信息收集表.xlsx")
session=openpyxl.load_workbook("./ApacheConSessions.xlsx")

# Getting data in a collection
collection = DOMTree.documentElement
twoCellAnchors = collection.getElementsByTagName("xdr"+":"+"twoCellAnchor")

# Take the string between two fixed symbols
def get_str_btw(str,begin,end):
   par=str.partition(begin)
   return (par[2].partition(end))[0][:]

# Get row and pic
for twoCellAnchor in twoCellAnchors:
 
   # Get row
   from1 = twoCellAnchor.getElementsByTagName('xdr'+':'+'to')[0]
   row1=from1.getElementsByTagName('xdr'+':'+'row')[0]
   row=int(row1.childNodes[0].data)

   # Get pic
   pic = twoCellAnchor.getElementsByTagName('xdr'+':'+'pic')[0]
   blipFill = pic.getElementsByTagName('xdr'+':'+'blipFill')[0]
   blip = blipFill.getElementsByTagName('a'+':'+'blip')[0]
   if blip.hasAttribute("r"+":"+"embed"):
      pic="picture"+(blip.getAttribute("r"+":"+"embed"))[3:]

   # Gets the active table object
   collect_active=collect.active
   session_active=session.active

   # Get the name, position, track and email address of the lecturer
   zh_name=collect_active.cell(row,3).value
   position=collect_active.cell(row,6).value
   track=collect_active.cell(row,9).value
   mail=collect_active.cell(row,10).value

   # Judge whether it is a Chinese name
   if '\u4e00' <=zh_name <='\u9fff':
      # Translate Chinese name into English name
      name_list=lazy_pinyin(zh_name)
      xin=name_list[0]
      ming_list=name_list[1:]
      ming=""
      for i in ming_list:
         ming=ming+i
      en_name=ming.capitalize()+"_"+xin.capitalize()
   else:
      en_name=zh_name

   # According to the mail from the collection table, find the corresponding row from the session table. 
   # If there is no corresponding row, return 0
   flag=False
   for cell1 in session_active['J']:
      str=cell1.value
      mail_list=str.split(",")
      for i in mail_list:
         session_mail=get_str_btw(i,'<','>')
         if session_mail==mail :
            flag=True
            session_row=cell1.row
            title=session_active.cell(cell1.row,7).value
            break
   if flag==False:
      session_row=0
      title=collect_active.cell(row,4).value       

   # Write data to a CSV file
   data=[
      (row,pic,zh_name,en_name,position,track,title,mail,session_row)
   ]
   with open("./mapping.csv","a+",encoding="utf-8",newline="") as f:
      csvf=csv.writer(f)
      csvf.writerows(data)
